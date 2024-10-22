#!/trinity/local/python/bin/python3
# -*- coding: utf-8 -*-
#pylint: disable=line-too-long,no-member,broad-except

# This code is part of the TrinityX software suite
# Copyright (C) 2023  ClusterVision Solutions b.v.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>


"""
limport Utility for Trinity Project
"""
__author__      = "Diego Sonaglia"
__copyright__   = "Copyright 2022, Luna2 Project [UTILITY]"
__license__     = "GPL"
__version__     = "2.0"
__maintainer__  = "Diego Sonaglia"
__email__       = "diego.sonaglia@clustervision.com"
__status__      = "Development"

import os
import re
import sys
import argparse
import ipaddress
import subprocess
from functools import partial
from itertools import zip_longest
from configparser import ConfigParser

import jwt
import openpyxl
import requests
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


TOKEN = None
LUNA_CONFIG_PATH = '/trinity/local/luna/utils/config/luna.ini'
settings = ConfigParser()
settings.read(LUNA_CONFIG_PATH)
requests.packages.urllib3.disable_warnings()


def print_info(msg):
    """Print an info message"""
    print(f'\033[1mINFO\033[0m: {msg}')

def print_error(msg):
    """Print an error message"""
    print(f'\033[91mERROR\033[0m: {msg}')

def print_fatal(msg):
    """Print a fatal error message"""
    print(f'\033[91mFATAL\033[0m: {msg}')


def get_token():
    """

    This method will fetch a valid token for further use.

    """
    # If there is a token check that is valid and return it
    if TOKEN is not None:
        try:
            # Try to decode the token to check if it is still valid
            jwt.decode(TOKEN, settings['API']['SECRET_KEY'], algorithms=['HS256'])
            return TOKEN
        except jwt.exceptions.ExpiredSignatureError:
            # If the token is expired is ok, we fetch a new one
            pass

    # Otherwise just fetch a new one
    data = {'username': settings['API']['USERNAME'], 'password': settings['API']['PASSWORD']}
    daemon_url = f"{settings['API']['PROTOCOL']}://{settings['API']['ENDPOINT']}/token"
    response = requests.post(
        daemon_url,
        json=data,
        stream=True,
        timeout=3,
        verify=(settings['API']['VERIFY_CERTIFICATE'].lower() == 'true'))
    token = response.json()['token']
    return token


def get_network_info(network_name):
    '''
    Get the network info from luna

    Args:
        network_name (str): The network name

    Returns:
        dict: The network info'''
    headers = {'x-access-tokens': get_token()}
    resp = requests.get(f"{settings['API']['PROTOCOL']}://{settings['API']['ENDPOINT']}/config/network/{network_name}",
                        stream=True,
                        timeout=3,
                        verify=(settings['API']['VERIFY_CERTIFICATE'].lower() == 'true'),
                        headers=headers)
    if resp.status_code != 200:
        return None

    return resp.json()['config']['network'][network_name]

def get_group_info(group_name):
    '''
    Get the group info from luna

    Args:
        group_name (str): The group name

    Returns:
        dict: The group info
    '''
    headers = {'x-access-tokens': get_token()}
    resp = requests.get(f"{settings['API']['PROTOCOL']}://{settings['API']['ENDPOINT']}/config/group/{group_name}",
                        stream=True,
                        timeout=3,
                        verify=(settings['API']['VERIFY_CERTIFICATE'].lower() == 'true'),
                        headers=headers)
    if resp.status_code != 200:
        return None

    return resp.json()['config']['group'][group_name]

def add_node(nodename, nodeinfo):
    '''
    Add the node to luna

    Args:
        nodename (str): The node name
        nodeinfo (dict): The node info

    Returns:
        int: The response status code
    '''

    headers = {'x-access-tokens': get_token()}
    resp = requests.post(f"{settings['API']['PROTOCOL']}://{settings['API']['ENDPOINT']}/config/node/{nodename}",
                         stream=True,
                         timeout=3,
                         verify=(settings['API']['VERIFY_CERTIFICATE'].lower() == 'true'),
                         json={'config':{'node': {nodename: nodeinfo}}},
                         headers=headers)
    return resp.status_code

class ExcelSheet():
    """
    Class that holds a generic excel sheet
    """
    def __init__(self, excel_path, sheet_name) -> None:
        self.excel_path = excel_path
        self.workbook =  self._open_excel()
        self.worksheet = self._open_worksheet(sheet_name)

    def _open_worksheet(self, sheet_name):
        try:
            return self.workbook[sheet_name]
        except Exception as exc:
            print_error(f"Cannot open the worksheet: {exc}")
            sys.exit(1)

    def _open_excel(self):
        try:
            return openpyxl.load_workbook(self.excel_path)
        except Exception as exc:
            print_error(f"Cannot open the excel file: {exc}")
            sys.exit(1)

    def _save_excel(self):
        try:
            self.workbook.save(self.excel_path)
        except Exception as exc:
            print_error(f"Cannot save the excel file: {exc}")
            sys.exit(1)

class NetworksSheet(ExcelSheet):
    """
    Class that holds the list of networks to be imported
    """
    header = [
        "boot_network_name",
        # "boot_network_addr",
        "ib_network_name",
        # "ib_network_addr",
        "ipmi_network_name",
        # "ipmi_network_addr",
        "group_name",
    ]

    def __init__(self, excel_path) -> None:
        super().__init__(excel_path, "Networks")

        self.validators = {
            "boot_network_name": self._validate_network_name,
            # "boot_network_addr": self._validate_network_addr,
            "ib_network_name": self._validate_network_name,
            # "ib_network_addr": self._validate_network_addr,
            "ipmi_network_name": self._validate_optional_network_name,
            # "ipmi_network_addr": self._validate_optional_network_addr,
            "group_name": self._validate_group_name,
        }

    @classmethod
    def _validate_network_name(cls, network_name):
        """Validate the network name

        Args:
            network_name (str): The network name to be validated

        Raises:
            ValueError: If the network name is not valid
        """
        if not network_name:
            raise ValueError("Network name cannot be empty")
        if not network_name.isalnum():
            raise ValueError(f"Network '{network_name}' contains invalid characters")

        network_info = get_network_info(network_name)
        if not network_info:
            raise ValueError(f"Network '{network_name}' not found")

    @classmethod
    def _validate_network_addr(cls, network_addr):
        """Validate the network address

        Args:
            network_addr (str): The network address to be validated

        Raises:
            ValueError: If the network address is not valid
        """
        if not network_addr:
            raise ValueError("Network address cannot be empty")
        if not ipaddress.ip_network(network_addr):
            raise ValueError("Invalid network address")
        if ipaddress.ip_network(network_addr).prefixlen == 32:
            raise ValueError("Network address cannot be a host address")

    @classmethod
    def _validate_optional_network_name(cls, network_name):
        """Validate the network name

        Args:
            network_name (str): The network name to be validated

        Raises:
            ValueError: If the network name is not valid
        """
        if not network_name:
            return
        cls._validate_network_name(network_name)

    @classmethod
    def _validate_optional_network_addr(cls, network_addr):
        """Validate the network address

        Args:
            network_addr (str): The network address to be validated

        Raises:
            ValueError: If the network address is not valid
        """
        if not network_addr:
            return
        cls._validate_network_addr(network_addr)

    @classmethod
    def _validate_group_name(cls, group_name):
        """Validate the group name

        Args:
            group_name (str): The group name to be validated

        Raises:
            ValueError: If the group name is not valid
        """
        if not group_name:
            raise ValueError("Group name cannot be empty")
        if not all(c in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_" for c in group_name):
            raise ValueError(f"Group '{group_name}' contains invalid characters")

        group_info = get_group_info(group_name)
        if not group_info:
            raise ValueError(f"Group '{group_name}' not found")

    def _validate_header(self):
        """Validate the header of the excel file

        Raises:
            ValueError: If the header is not valid
        """
        for row, expected_cell in zip_longest(self.worksheet.iter_rows(), self.header):
            cell = row[0]
            if cell.value != expected_cell:
                raise ValueError(f"Invalid header in cell {cell.coordinate}: {cell.value}")

    def _validate_networks(self):
        """Validate the networks in the excel file

        Returns:
            bool: True if there is an error, False otherwise
        """
        is_error = False
        for row in self.worksheet.iter_rows(min_row=1):
            header_cell = row[0]
            cell = row[1]
            validator = self.validators[header_cell.value]
            try:
                validator(cell.value)
                cell.fill = PatternFill(start_color="C6EFCE",
                                        end_color="C6EFCE",
                                        fill_type = "solid")
                cell.comment = None
            except (ValueError, TypeError, AttributeError) as exc:
                print_error(f"Invalid value in cell {cell.coordinate}: {exc}")
                cell.fill = PatternFill(start_color="FFC7CE",
                                        end_color="FFC7CE",
                                        fill_type = "solid")
                cell.comment = Comment(str(exc), "Invalid")
                is_error = True
        return is_error

    def validate(self):
        """Validate the excel file

        Raises:
            ValueError: If the excel file is not valid
        """
        print_info("Validating Networks sheet")
        self._validate_header()
        is_error = self._validate_networks()
        self._save_excel()
        if is_error:
            raise ValueError("Invalid excel file")

        print_info("Networks sheet ok!")

    def get_networks_info(self):
        """Get the networks info from the excel file"""
        return {
            'boot_network': get_network_info(self.worksheet['B1'].value),
            'ib_network'  : get_network_info(self.worksheet['B2'].value),
            'ipmi_network': get_network_info(self.worksheet['B3'].value),
            'group': get_group_info(self.worksheet['B4'].value),
        }

class NodesSheet(ExcelSheet):
    """
    Class that holds the list of nodes to be imported
    """
    header = [
        "hostname",
        "mac_addr",
        "ip_addr_bootif",
        "ip_addr_ib",
        "ip_addr_ipmi"
    ]

    def __init__(self, excel_path, networks_info) -> None:
        super().__init__(excel_path, "Nodes")

        self.networks_info = networks_info
        self.validators = {
            "hostname": self._validate_hostname,
            "mac_addr": self._validate_mac_address,
            "ip_addr_bootif": partial(self._validate_ip_address, network_info=self.networks_info['boot_network']),
            "ip_addr_ib": partial(self._validate_optional_ip_address, network_info=self.networks_info['ib_network']),
            "ip_addr_ipmi": partial(self._validate_optional_ip_address, network_info=self.networks_info['ipmi_network']),
        }

    @classmethod
    def _validate_hostname(cls, hostname):
        """Validate the hostname

        Args:
            hostname (str): The hostname to be validated

        Raises:
            ValueError: If the hostname is not valid
        """
        if not hostname:
            raise ValueError("Hostname cannot be empty")
        if not hostname.isalnum():
            raise ValueError("Hostname contains invalid characters")

    @classmethod
    def _validate_mac_address(cls, mac_address):
        """Validate the mac address

        Args:
            mac_address (str): The mac address to be validated

        Raises:
            ValueError: If the mac address is not valid
        """
        if not mac_address:
            return
        if not re.match(r"([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})", mac_address):
            raise ValueError("Invalid mac address")

    @classmethod
    def _validate_ip_address(cls, ip_address, network_info):
        """Validate the ip address

        Args:
            ip_address (str): The ip address to be validated

        Raises:
            ValueError: If the ip address is not set or not valid
        """
        if not ipaddress.ip_address(ip_address):
            raise ValueError("Invalid ip address")
        if not network_info:
            raise ValueError("Network does not exist")

        network_addr = ipaddress.ip_network(network_info['network'])
        node_addr = ipaddress.ip_address(ip_address)
        if node_addr not in network_addr:
            raise ValueError(f"IP address {ip_address} is not in the network {network_info['network']}")

    @classmethod
    def _validate_optional_ip_address(cls, ip_address, network_info):
        """Validate the ip address

        Args:
            ip_address (str): The ip address to be validated

        Raises:
            ValueError: If the ip address is not valid
        """
        if not ip_address:
            return
        cls._validate_ip_address(ip_address, network_info)

    def _validate_header(self):
        """Validate the header of the excel file

        Raises:
            ValueError: If the header is not valid
        """
        header = self.worksheet[1]
        for _, (col, expected_col) in enumerate(zip_longest(header, self.header),1):
            if col.value != expected_col:
                raise ValueError(f"Invalid header in cell {col.coordinate}. expected: {header} got: {col.value}")

    def _validate_nodes(self):
        """Validate the nodes of the excel file

        Returns:
            bool: True if there is an error, False otherwise
        """
        is_error = False
        nodes = self.worksheet
        for row in nodes.iter_rows(min_row=2):
            for col_inedx, col in enumerate(self.header, 1):
                cell = row[col_inedx-1]
                validator = self.validators[col]
                try:
                    validator(cell.value)
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type = "solid")
                    cell.comment = None
                except (ValueError, TypeError, AttributeError) as exc:
                    print_error(f"Invalid value in cell {cell.coordinate}: {exc}")
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                    cell.comment = Comment(str(exc), "Invalid")
                    is_error = True
        return is_error

    def validate(self):
        """Validate the excel file

        Raises:
            ValueError: If the excel file is not valid
        """
        print_info("Validating Nodes sheet")
        self._validate_header()
        is_error = self._validate_nodes()
        self._save_excel()
        if is_error:
            raise ValueError("Invalid excel file")

        print_info("Nodes sheet ok!")

    def get_nodes(self):
        """Get the nodes from the excel file

        Returns:
            list: The list of nodes
        """
        nodes = []
        for row in self.worksheet.iter_rows(min_row=2):
            node = {}
            node['name'] = row[0].value
            nodename = row[0].value
            node['group'] = self.networks_info['group']['name']
            node['interfaces'] = []
            node['interfaces'].append({
                'interface': 'BOOTIF',
                'macaddress': row[1].value,
                'ipaddress': row[2].value,
                'network': self.networks_info['boot_network']['name']
            })
            if row[3].value:
                node['interfaces'].append({
                    'interface': 'IB0',
                    'ipaddress': row[3].value,
                    'network': self.networks_info['ib_network']['name']
                })
            if row[4].value:
                node['interfaces'].append({
                    'interface': 'BMC',
                    'ipaddress': row[4].value,
                    'network': self.networks_info['ipmi_network']['name']
                })
            nodes.append((nodename, node))
        return nodes

    def add_nodes(self):
        '''
        Add the nodes to luna
        '''
        nodes = self.get_nodes()
        for nodename, nodeinfo in nodes:
            status_code = add_node(nodename, nodeinfo)
            if status_code == 201:
                print_info(f"Node '{nodename}' added successfully")
            elif status_code == 204:
                print_info(f"Node '{nodename}' modified successfully")
            else:
                print_fatal(f"Cannot add node '{nodename}', returned status_code: {status_code}")
                sys.exit(1)

class CLI():
    """
    Command that allows to bulk import nodes to luna from an excel file
    """
    @classmethod
    def add(cls, excel_path: str):
        """Check the excel file to be imported and add the nodes to luna

        Args:
            excel_path (str): The excel file to be imported
        """
        try:
            network_sheet = NetworksSheet(excel_path)
            network_sheet.validate()
            nodes_sheet = NodesSheet(
                excel_path,
                networks_info=network_sheet.get_networks_info()
                )
            nodes_sheet.validate()
            nodes_sheet.add_nodes()

        except ValueError as exc:
            print_error(exc)
            sys.exit(1)

    @classmethod
    def check(cls, excel_path: str):
        """Check the excel file to be imported

        Args:
            excel_path (str): The excel file to be checked"""
        try:
            network_sheet = NetworksSheet(excel_path)
            network_sheet.validate()
            nodes_sheet = NodesSheet(
                excel_path,
                networks_info=network_sheet.get_networks_info()
                )
            nodes_sheet.validate()
        except ValueError as exc:
            print_error(exc)
            sys.exit(1)

    @classmethod
    def copy_template(cls, target_path):
        """Copy the reference excel template to the target path

        Args:
            target_path (str): The path where the excel file will be generated
        """
        template_path = os.path.join(
            os.path.dirname(os.path.realpath(__file__)),
            '..',
            'data',
            'limport',
            'nodes.xlsx')
        try:
            subprocess.check_output(['cp', template_path, target_path])
            print_info(f"Template successfully copied in {target_path}")
        except Exception as exc:
            print_error(f"Cannot copy the template to {target_path}: {exc}")
            sys.exit(1)


def main():
    '''
    Main function
    '''
    parser = argparse.ArgumentParser(description='Utility to import nodes in bulk to luna')
    subparsers = parser.add_subparsers(help='sub-command help', dest='command')

    parser_add = subparsers.add_parser('add', help='Add nodes to luna')
    parser_add.add_argument('excel_path', help='The excel file to be imported')

    parser_check = subparsers.add_parser('check', help='check the excel file')
    parser_check.add_argument('excel_path', help='The excel file to be checkd')

    parser_template = subparsers.add_parser('copy_template', help='Copy the reference excel template')
    parser_template.add_argument('target_path', help='The path where the excel file will be generated')

    args = parser.parse_args()
    if args.command == 'add':
        CLI.add(args.excel_path)
    elif args.command == 'check':
        CLI.check(args.excel_path)
    elif args.command == 'copy_template':
        CLI.copy_template(args.target_path)
    else:
        parser.print_help()



if __name__ == '__main__':
    main()
